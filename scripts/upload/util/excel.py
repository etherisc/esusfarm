from loguru import logger

from openpyxl import load_workbook

EXCEL_ROW_ID = 'excel_row_num'


def validate_column_headers(name, sheet, header, exception_type, strict_header_order):
    if strict_header_order:
        for (label, column) in header.items():
            cell = "{0}1".format(column)
            if label != sheet[cell].value.strip():
                raise ExcelException("{exception} exception. sheet: {sheet}, cell: {cell}, expected: '{expected}', actual: '{actual}'".format(
                    exception=exception_type, sheet=name, cell=cell, expected=label, actual=sheet[cell].value
                ))

        return header
    # check that all header attributes are available (and available exactly once)
    else:
        # prepare attribute check
        attributes = {}
        header_found = {}
        for attribute in header:
            attributes[attribute] = 0

        # run over all column labels (and remember which column label has which column coordinates)
        for col in sheet.iter_cols(min_row=1, max_row=1, max_col=sheet.max_column):
            for cell in col:
                logger.debug("excel header cell {} row {} column {}/{} value {}".format(cell, cell.row, cell.column, cell.column_letter, cell.value))

                if cell.value:
                    attribute_name = cell.value.strip()
                    header_found[attribute_name] = cell.column_letter

                    if attribute_name in attributes:
                        attributes[attribute_name] += 1
        
        # assess the outcome
        last_exception = None

        for (attribute, count) in attributes.items():
            message = None

            if count == 0:
                message = "{exception} exception. sheet: {sheet}, missing attribute: '{attribute}'".format(
                    exception=exception_type, sheet=name, attribute=attribute)
            elif count > 1:
                message = "{exception} exception. sheet: {sheet}, attribute '{attribute}' occurs {count} times (expected only once)".format(
                    exception=exception_type, sheet=name, attribute=attribute, count=count)
            
            if message:
                logger.warning(message)
                last_exception = ExcelException(message)
        
        if last_exception:
            raise last_exception
        
        # remove unnecessary header mappings
        header_out = {}

        for attribute in header_found:
            if attribute in attributes:
                header_out[attribute] = header_found[attribute]
        
        return header_out


def extract_row_data(sheet, header, row, strict_header_order):
    row_data = {}
    row_data[EXCEL_ROW_ID] = row

    for (attribute, column) in header.items():
        cell = "{}{}".format(column, row)
        value = sheet[cell].value

        if isinstance(value, str):
            row_data[attribute] = value.strip()
        else:
            row_data[attribute] = value

    return row_data


def get_data(excel_file_path, header, strict_header_order=True, sheet_names=None):
    logger.info("loading excel file '{}' ...".format(excel_file_path))

    wb = load_workbook(excel_file_path, data_only=True)
    data = {}

    # run through all availble sheets
    for sheet_name in wb.sheetnames:
        if sheet_names and sheet_name not in sheet_names:
            logger.info("invalid sheet name '{}' skipping ...".format(sheet_name))
            continue

        logger.info("processing sheet '{}' ...".format(sheet_name))

        data[sheet_name] = {}
        sheet = wb[sheet_name]

        # validate header row
        try:
            header_out = validate_column_headers(sheet_name, sheet, header, excel_file_path, strict_header_order)
        except Exception as e:
            raise ExcelException("header column validation error for sheet '{}': {}".format(sheet_name, e))

        # extract cell values from each row
        row = 2
        while sheet["A{}".format(row)].value:
            data[sheet_name][row] = extract_row_data(sheet, header_out, row, strict_header_order)
            row += 1
        
    return data


class ExcelException(Exception):

    def __init__(self, message):
        self.err = message
        super().__init__(self.err)
